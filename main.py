import openpyxl
import xml.etree.ElementTree as orgs
# generates root for xml file
root = orgs.Element("organization_units")
root.set('xsi:noNamespaceSchemaLocation', 'organization_units.xsd')
root.set('xmlns:xsi', 'http://www.w3.org/2001/XMLSchema-instance')

# generates a subunit for the xml file
def GenerateXMLRecord(dName, dCode, fieldToAppendTo, type):
    unitData = orgs.Element("unitData")
    fieldToAppendTo.append(unitData)
    o1 = orgs.SubElement(unitData, "organizationCode")
    o1.text = dCode
    o2 = orgs.SubElement(unitData, "organizationName")
    o2.text = dName
    o3 = orgs.SubElement(unitData, "unitType")
    o3.text = type

if __name__ == '__main__':
    filepath = "/Users/library/Desktop/collegeOrgs.xlsx"
    wrkbk = openpyxl.load_workbook(filepath)
    sh = wrkbk.active
    # array of colleges to look for in excel file
    colleges = ["Academic and Student Affairs", "Libraries and Learning Commons", "College of Business and Public Management", "LaFetra College of Education", "College of Arts and Sciences", "College of Law", "President's Office", "University Advancement"]
    # iterates through excel file and searches for each college name
    for j in range(0, len(colleges)):
        collegeName = colleges[j]
        collegeCode = collegeName.replace(" ", "-").lower()
        # generates a unit group for each college
        mainUnit = orgs.Element("unit")
        root.append(mainUnit)
        GenerateXMLRecord(collegeName, collegeCode, mainUnit, "esploro.organization.unit.types.college")
        subunits = orgs.Element("subUnits")
        mainUnit.append(subunits)
        #iterates through all rows in excel file to find specified college departments
        for i in range(1, sh.max_row + 1):
            excelName = sh.cell(row=i, column=1).value
            if excelName == collegeName:
                # generates subunits for each department in the current college
                departmentName = sh.cell(row=i, column=3).value
                departmentCode = sh.cell(row=i, column=4).value
                unit = orgs.Element("unit")
                subunits.append(unit)
                GenerateXMLRecord(departmentName, departmentCode, unit, "esploro.organization.unit.types.department")

    #  writes xml records to file 'organization_units.xml'
    tree = orgs.ElementTree(root)
    with open("organization_units.xml", "wb") as files:
        tree.write(files, "UTF-8", True)
    print("\nXML Conversion Complete\nGoodbye!")
